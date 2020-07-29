#!/usr/bin/env python
# coding=utf-8
import shutil
import openpyxl
from config import CF
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from common.variables import VariablePool


class ExcelSet:
    """Excel配置"""

    def __init__(self):
        shutil.copyfile(VariablePool.get('excel_input'), VariablePool.get('excel_output'))
        self.path = VariablePool.get('excel_output')
        self.wb = openpyxl.load_workbook(self.path)
        self.table = self.wb.active

    def get_cases(self, min_row=2):
        """获取用例"""
        all_cases = []
        for row in self.table.iter_rows(min_row=min_row):
            all_cases.append((self.table.cell(min_row, CF.NAME + 1).value,
                              min_row, [cell.value for cell in row]))
            min_row += 1
        return all_cases

    def write_color(self, row_n, col_n, fgColor="FF0000"):
        """写入颜色"""
        cell = self.table.cell(row_n, col_n + 1)
        fill = PatternFill("solid", fgColor=fgColor)
        cell.fill = fill

    def write_results(self, row_n, col_n, value, color=True):
        """写入结果"""
        cell = self.table.cell(row_n, col_n + 1)
        cell.value = value
        font = Font(name='微软雅黑', size=16)
        cell.font = font
        if color:
            if value.lower() in ("fail", 'failed'):
                fill = PatternFill("solid", fgColor="FF0000")
                cell.fill = fill
            elif value.lower() in ("pass", "ok"):
                fill = PatternFill("solid", fgColor="00CD00")
                cell.fill = fill
        self.wb.save(self.path)


excel_set = ExcelSet()
if __name__ == '__main__':
    print(excel_set.get_cases())
